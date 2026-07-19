from __future__ import annotations

import asyncio
import logging
import os
import sqlite3
from datetime import date, datetime, time, timedelta
from html import escape
from io import BytesIO
from zoneinfo import ZoneInfo

from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BotCommand,
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
)
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db import Challenge, Database


from pathlib import Path
dotenv_path = Path(__file__).parent / ".env"
load_dotenv(dotenv_path, override=False)

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
DATABASE_PATH = os.getenv("DATABASE_PATH", "sport_challenge.db")
TIMEZONE = ZoneInfo(os.getenv("TIMEZONE", "Europe/Moscow"))
RESULTS_CHAT_ID_RAW = os.getenv("RESULTS_CHAT_ID", "").strip()
RESULTS_CHAT_ID = int(RESULTS_CHAT_ID_RAW) if RESULTS_CHAT_ID_RAW else None
REMINDER_TIME = time.fromisoformat(os.getenv("REMINDER_TIME", "20:00"))
YESTERDAY_EDIT_UNTIL = time.fromisoformat(
    os.getenv("YESTERDAY_EDIT_UNTIL", "12:00")
)
ADMIN_IDS = {
    int(value.strip())
    for value in os.getenv("ADMIN_IDS", "").split(",")
    if value.strip()
}

if not BOT_TOKEN:
    raise RuntimeError("В .env не задан BOT_TOKEN")

db = Database(DATABASE_PATH)
router = Router()


class ResultForm(StatesGroup):
    pushups = State()
    pullups = State()
    squats = State()
    confirmation = State()


class ChallengeForm(StatesGroup):
    title = State()
    start_date = State()
    duration = State()


def now_local() -> datetime:
    return datetime.now(TIMEZONE)


def today_iso() -> str:
    return now_local().date().isoformat()


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def main_keyboard(admin: bool = False) -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(text="📝 Внести результат")],
        [
            KeyboardButton(text="📊 Моя статистика"),
            KeyboardButton(text="🏆 Общий рейтинг"),
        ],
        [
            KeyboardButton(text="📅 Результаты по дням"),
            KeyboardButton(text="🏅 Рейтинг недели"),
        ],
        [
            KeyboardButton(text="🔔 Напоминания"),
            KeyboardButton(text="ℹ️ Правила"),
        ],
    ]
    if admin:
        rows.append([KeyboardButton(text="🛠 Админ-панель")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)


def admin_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="➕ Новый челлендж"),
                KeyboardButton(text="🏁 Завершить"),
            ],
            [
                KeyboardButton(text="👀 Кто не внёс"),
                KeyboardButton(text="📥 Excel"),
            ],
            [
                KeyboardButton(text="🏅 Недельный рейтинг"),
                KeyboardButton(text="🏆 Итоговые номинации"),
            ],
            [KeyboardButton(text="⬅️ Главное меню")],
        ],
        resize_keyboard=True,
    )


def date_choice_keyboard(allow_yesterday: bool) -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(text="Сегодня", callback_data="result_date:today")]
    ]
    if allow_yesterday:
        buttons.append(
            [InlineKeyboardButton(text="Вчера", callback_data="result_date:yesterday")]
        )
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def confirmation_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Сохранить", callback_data="result:save"),
                InlineKeyboardButton(text="✏️ Исправить", callback_data="result:edit"),
            ],
            [InlineKeyboardButton(text="❌ Отмена", callback_data="result:cancel")],
        ]
    )


def challenge_available_for_date(
    challenge: Challenge, result_date: date
) -> tuple[bool, str]:
    start = date.fromisoformat(challenge.start_date)
    end = date.fromisoformat(challenge.end_date)
    if result_date < start:
        return False, "На эту дату челлендж ещё не начался."
    if result_date > end:
        return False, "Эта дата находится за пределами челленджа."
    return True, ""


def parse_number(message: Message, limit: int) -> int | None:
    text_value = (message.text or "").strip()
    if not text_value.isdigit():
        return None
    value = int(text_value)
    return value if 0 <= value <= limit else None


def weekly_range(challenge: Challenge, reference: date | None = None) -> tuple[int, date, date]:
    start = date.fromisoformat(challenge.start_date)
    end = date.fromisoformat(challenge.end_date)
    ref = min(max(reference or now_local().date(), start), end)
    week_number = ((ref - start).days // 7) + 1
    week_start = start + timedelta(days=(week_number - 1) * 7)
    week_end = min(week_start + timedelta(days=6), end)
    return week_number, week_start, week_end


def build_ranking_text(
    challenge: Challenge,
    date_from: str | None = None,
    date_to: str | None = None,
    heading: str | None = None,
) -> str:
    rows = db.get_ranking(challenge.id, date_from, date_to)
    title = heading or f"🏆 {challenge.title}"
    if not rows:
        return f"<b>{escape(title)}</b>\n\nПока результатов нет."

    lines = [f"<b>{escape(title)}</b>", ""]
    medals = ["🥇", "🥈", "🥉"]
    for index, row in enumerate(rows, start=1):
        marker = medals[index - 1] if index <= 3 else f"{index}."
        lines.append(
            f"{marker} <b>{escape(row['full_name'])}</b> — "
            f"{row['points']:.2f}\n"
            f"   Дней: {row['days']} · Отж.: {row['pushups']} · "
            f"Подт.: {row['pullups']} · Присед.: {row['squats']}"
        )
    return "\n".join(lines)


def build_weekly_text(challenge: Challenge, reference: date | None = None) -> str:
    number, start, end = weekly_range(challenge, reference)
    return build_ranking_text(
        challenge,
        start.isoformat(),
        end.isoformat(),
        f"🏅 Неделя {number}: {start:%d.%m}–{end:%d.%m}",
    )


def build_daily_text(challenge: Challenge, result_date: date) -> str:
    rows = db.get_daily_results(challenge.id, result_date.isoformat())
    missing = db.get_missing_users(challenge.id, result_date.isoformat())
    lines = [f"📅 <b>{result_date:%d.%m.%Y}</b>", ""]
    if rows:
        for row in rows:
            lines.append(
                f"✅ <b>{escape(row['full_name'])}</b>: "
                f"{row['pushups']} / {row['pullups']} / {row['squats']} "
                f"— {row['points']:.2f}"
            )
    else:
        lines.append("Результатов пока нет.")
    if missing:
        lines.extend(["", "⏳ <b>Не внесли:</b>"])
        lines.extend(f"• {escape(row['full_name'])}" for row in missing)
    return "\n".join(lines)


def build_nominations_text(challenge: Challenge) -> str:
    rows = db.get_ranking(challenge.id)
    if not rows:
        return "🏆 Пока недостаточно результатов для номинаций."

    def best(field: str):
        return sorted(rows, key=lambda row: (-row[field], row["full_name"]))[0]

    overall = rows[0]
    pushups = best("pushups")
    pullups = best("pullups")
    squats = best("squats")
    perfect = best("perfect_days")

    longest = []
    for row in rows:
        _, max_streak = db.calculate_streaks(
            challenge.id, row["telegram_id"], challenge.end_date
        )
        longest.append((max_streak, row["full_name"]))
    streak_days, streak_name = sorted(longest, key=lambda item: (-item[0], item[1]))[0]

    return (
        f"🏆 <b>Итоговые номинации</b>\n\n"
        f"🥇 Общий победитель — <b>{escape(overall['full_name'])}</b> "
        f"({overall['points']:.2f})\n"
        f"💪 Отжимания — <b>{escape(pushups['full_name'])}</b> "
        f"({pushups['pushups']})\n"
        f"🦍 Подтягивания — <b>{escape(pullups['full_name'])}</b> "
        f"({pullups['pullups']})\n"
        f"🦵 Приседания — <b>{escape(squats['full_name'])}</b> "
        f"({squats['squats']})\n"
        f"⭐ Идеальные дни — <b>{escape(perfect['full_name'])}</b> "
        f"({perfect['perfect_days']})\n"
        f"🔥 Самая длинная серия — <b>{escape(streak_name)}</b> "
        f"({streak_days} дн.)"
    )


def build_final_text(challenge: Challenge) -> str:
    return (
        f"🏁 <b>Челлендж «{escape(challenge.title)}» завершён!</b>\n\n"
        f"{build_nominations_text(challenge)}\n\n"
        f"{build_ranking_text(challenge)}"
    )


def create_excel(challenge: Challenge) -> bytes:
    wb = Workbook()
    ws_rating = wb.active
    ws_rating.title = "Общий рейтинг"
    ws_days = wb.create_sheet("Результаты по дням")
    ws_users = wb.create_sheet("Участники")

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)

    rating_headers = [
        "Место", "Участник", "Баллы", "Дней", "Идеальных дней",
        "Отжимания", "Подтягивания", "Приседания",
        "Текущая серия", "Максимальная серия",
    ]
    ws_rating.append(rating_headers)
    for cell in ws_rating[1]:
        cell.fill = header_fill
        cell.font = header_font

    for index, row in enumerate(db.get_ranking(challenge.id), start=1):
        current_streak, max_streak = db.calculate_streaks(
            challenge.id, row["telegram_id"], min(today_iso(), challenge.end_date)
        )
        ws_rating.append([
            index, row["full_name"], round(row["points"], 2), row["days"],
            row["perfect_days"], row["pushups"], row["pullups"], row["squats"],
            current_streak, max_streak,
        ])

    day_headers = [
        "Дата", "Участник", "Username", "Отжимания", "Подтягивания",
        "Приседания", "Баллы", "Последнее изменение",
    ]
    ws_days.append(day_headers)
    for cell in ws_days[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in db.get_all_results(challenge.id):
        ws_days.append([
            row["result_date"], row["full_name"], row["username"] or "",
            row["pushups"], row["pullups"], row["squats"],
            round(row["points"], 2), row["updated_at"],
        ])

    ws_users.append(["Участник", "Username", "Напоминания"])
    for cell in ws_users[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in db.get_all_users():
        ws_users.append([
            row["full_name"], row["username"] or "",
            "Включены" if row["reminders_enabled"] else "Выключены",
        ])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_len + 2, 11), 32)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@router.message(CommandStart())
async def start(message: Message, state: FSMContext) -> None:
    await state.clear()
    if not message.from_user:
        return
    db.upsert_user(
        message.from_user.id,
        message.from_user.full_name,
        message.from_user.username,
    )
    await message.answer(
        "Привет! Я веду спортивный челлендж.",
        reply_markup=main_keyboard(is_admin(message.from_user.id)),
    )


@router.message(Command("id"))
async def show_id(message: Message) -> None:
    user_id = message.from_user.id if message.from_user else 0
    await message.answer(
        f"Ваш Telegram ID: <code>{user_id}</code>\n"
        f"ID этого чата: <code>{message.chat.id}</code>"
    )


@router.message(F.text == "📝 Внести результат")
async def choose_result_date(message: Message, state: FSMContext) -> None:
    if not message.from_user:
        return
    db.upsert_user(
        message.from_user.id,
        message.from_user.full_name,
        message.from_user.username,
    )
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Сейчас нет активного челленджа.")
        return

    now = now_local()
    allow_yesterday = now.time() < YESTERDAY_EDIT_UNTIL
    await state.clear()
    await message.answer(
        "За какой день внести результат?",
        reply_markup=date_choice_keyboard(allow_yesterday),
    )


@router.callback_query(F.data.startswith("result_date:"))
async def select_result_date(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user:
        return
    challenge = db.get_active_challenge()
    if not challenge:
        await callback.answer("Нет активного челленджа.", show_alert=True)
        return

    selected = callback.data.split(":", 1)[1]
    target = now_local().date()
    if selected == "yesterday":
        if now_local().time() >= YESTERDAY_EDIT_UNTIL:
            await callback.answer("Срок ввода за вчера уже истёк.", show_alert=True)
            return
        target -= timedelta(days=1)

    available, reason = challenge_available_for_date(challenge, target)
    if not available:
        await callback.answer(reason, show_alert=True)
        return

    await state.set_state(ResultForm.pushups)
    await state.update_data(challenge_id=challenge.id, result_date=target.isoformat())
    old = db.get_result(challenge.id, callback.from_user.id, target.isoformat())
    old_text = ""
    if old:
        old_text = (
            f"\nУже записано: {old['pushups']} / {old['pullups']} / "
            f"{old['squats']}. Новые значения заменят старые."
        )
    if callback.message:
        await callback.message.answer(
            f"Сколько отжиманий за {target:%d.%m}? "
            f"Введите 0–{challenge.pushup_limit}.{old_text}",
            reply_markup=ReplyKeyboardRemove(),
        )
    await callback.answer()


@router.message(ResultForm.pushups)
async def input_pushups(message: Message, state: FSMContext) -> None:
    challenge = db.get_active_challenge()
    if not challenge:
        await state.clear()
        return
    value = parse_number(message, challenge.pushup_limit)
    if value is None:
        await message.answer(f"Введите целое число 0–{challenge.pushup_limit}.")
        return
    await state.update_data(pushups=value)
    await state.set_state(ResultForm.pullups)
    await message.answer(f"Подтягивания: 0–{challenge.pullup_limit}.")


@router.message(ResultForm.pullups)
async def input_pullups(message: Message, state: FSMContext) -> None:
    challenge = db.get_active_challenge()
    if not challenge:
        await state.clear()
        return
    value = parse_number(message, challenge.pullup_limit)
    if value is None:
        await message.answer(f"Введите целое число 0–{challenge.pullup_limit}.")
        return
    await state.update_data(pullups=value)
    await state.set_state(ResultForm.squats)
    await message.answer(f"Приседания: 0–{challenge.squat_limit}.")


@router.message(ResultForm.squats)
async def input_squats(message: Message, state: FSMContext) -> None:
    challenge = db.get_active_challenge()
    if not challenge:
        await state.clear()
        return
    value = parse_number(message, challenge.squat_limit)
    if value is None:
        await message.answer(f"Введите целое число 0–{challenge.squat_limit}.")
        return
    await state.update_data(squats=value)
    data = await state.get_data()
    points = (
        data["pushups"] / challenge.pushup_limit
        + data["pullups"] / challenge.pullup_limit
        + value / challenge.squat_limit
    )
    await state.set_state(ResultForm.confirmation)
    await message.answer(
        f"Проверьте результат за {date.fromisoformat(data['result_date']):%d.%m.%Y}:\n\n"
        f"Отжимания: <b>{data['pushups']}</b>\n"
        f"Подтягивания: <b>{data['pullups']}</b>\n"
        f"Приседания: <b>{value}</b>\n"
        f"Баллы: <b>{points:.2f}</b>",
        reply_markup=confirmation_keyboard(),
    )


@router.callback_query(ResultForm.confirmation, F.data == "result:save")
async def save_confirmed(callback: CallbackQuery, state: FSMContext) -> None:
    if not callback.from_user:
        return
    data = await state.get_data()
    challenge = db.get_active_challenge()
    if not challenge or challenge.id != data.get("challenge_id"):
        await state.clear()
        await callback.answer("Челлендж уже закрыт.", show_alert=True)
        return

    db.save_result(
        challenge.id,
        callback.from_user.id,
        data["result_date"],
        data["pushups"],
        data["pullups"],
        data["squats"],
    )
    current_streak, max_streak = db.calculate_streaks(
        challenge.id, callback.from_user.id, data["result_date"]
    )
    await state.clear()
    if callback.message:
        await callback.message.answer(
            f"✅ Результат сохранён.\n"
            f"🔥 Текущая серия: {current_streak} дн.\n"
            f"Лучшая серия: {max_streak} дн.",
            reply_markup=main_keyboard(is_admin(callback.from_user.id)),
        )
    await callback.answer()


@router.callback_query(ResultForm.confirmation, F.data == "result:edit")
async def edit_result(callback: CallbackQuery, state: FSMContext) -> None:
    await state.set_state(ResultForm.pushups)
    if callback.message:
        challenge = db.get_active_challenge()
        limit = challenge.pushup_limit if challenge else 200
        await callback.message.answer(
            f"Введите отжимания заново: 0–{limit}.",
            reply_markup=ReplyKeyboardRemove(),
        )
    await callback.answer()


@router.callback_query(ResultForm.confirmation, F.data == "result:cancel")
async def cancel_result_callback(callback: CallbackQuery, state: FSMContext) -> None:
    await state.clear()
    if callback.message:
        await callback.message.answer(
            "Ввод отменён.",
            reply_markup=main_keyboard(is_admin(callback.from_user.id)),
        )
    await callback.answer()


@router.message(Command("cancel"))
async def cancel(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "Действие отменено.",
        reply_markup=main_keyboard(bool(message.from_user and is_admin(message.from_user.id))),
    )


@router.message(F.text == "📊 Моя статистика")
async def my_stats(message: Message) -> None:
    if not message.from_user:
        return
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Сейчас нет активного челленджа.")
        return
    stats = db.get_user_stats(challenge.id, message.from_user.id)
    if not stats or stats["days"] == 0:
        await message.answer("У вас пока нет результатов.")
        return
    current_streak, max_streak = db.calculate_streaks(
        challenge.id, message.from_user.id, min(today_iso(), challenge.end_date)
    )
    await message.answer(
        f"📊 <b>Ваша статистика</b>\n\n"
        f"Баллы: <b>{stats['points']:.2f}</b>\n"
        f"Дней: {stats['days']}\n"
        f"Отжимания: {stats['pushups']}\n"
        f"Подтягивания: {stats['pullups']}\n"
        f"Приседания: {stats['squats']}\n"
        f"Лучший день: {stats['best_day']:.2f} балла\n"
        f"🔥 Текущая серия: {current_streak}\n"
        f"🏅 Лучшая серия: {max_streak}"
    )


@router.message(F.text == "🏆 Общий рейтинг")
@router.message(Command("rating"))
async def rating(message: Message) -> None:
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Сейчас нет активного челленджа.")
        return
    await message.answer(build_ranking_text(challenge))


@router.message(F.text.in_({"🏅 Рейтинг недели", "🏅 Недельный рейтинг"}))
async def weekly_rating(message: Message) -> None:
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Сейчас нет активного челленджа.")
        return
    await message.answer(build_weekly_text(challenge))


@router.message(F.text == "📅 Результаты по дням")
async def daily_results(message: Message) -> None:
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Сейчас нет активного челленджа.")
        return
    target = min(now_local().date(), date.fromisoformat(challenge.end_date))
    await message.answer(build_daily_text(challenge, target))


@router.message(F.text == "🔔 Напоминания")
async def toggle_reminders(message: Message) -> None:
    if not message.from_user:
        return
    db.upsert_user(
        message.from_user.id,
        message.from_user.full_name,
        message.from_user.username,
    )
    enabled = db.toggle_reminders(message.from_user.id)
    await message.answer(
        "🔔 Напоминания включены." if enabled else "🔕 Напоминания выключены."
    )


@router.message(F.text == "ℹ️ Правила")
async def rules(message: Message) -> None:
    await message.answer(
        "ℹ️ <b>Правила</b>\n\n"
        "Дневные максимумы: 200 отжиманий, 50 подтягиваний, "
        "200 приседаний.\n"
        "За каждое упражнение начисляется до 1 балла. "
        "Максимум — 3 балла в день.\n\n"
        f"За вчера можно внести результат до {YESTERDAY_EDIT_UNTIL:%H:%M}. "
        "Повторный ввод заменяет старое значение."
    )


@router.message(F.text == "🛠 Админ-панель")
async def admin_panel(message: Message) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await message.answer("Управление челленджем:", reply_markup=admin_keyboard())


@router.message(F.text == "⬅️ Главное меню")
async def main_menu(message: Message, state: FSMContext) -> None:
    await state.clear()
    await message.answer(
        "Главное меню.",
        reply_markup=main_keyboard(bool(message.from_user and is_admin(message.from_user.id))),
    )


@router.message(F.text == "➕ Новый челлендж")
async def new_challenge_start(message: Message, state: FSMContext) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    if db.get_active_challenge():
        await message.answer("Сначала завершите текущий челлендж.")
        return
    await state.set_state(ChallengeForm.title)
    await message.answer("Введите название челленджа:", reply_markup=ReplyKeyboardRemove())


@router.message(ChallengeForm.title)
async def new_challenge_title(message: Message, state: FSMContext) -> None:
    title = (message.text or "").strip()
    if not title:
        await message.answer("Название не должно быть пустым.")
        return
    await state.update_data(title=title)
    await state.set_state(ChallengeForm.start_date)
    await message.answer("Введите дату начала в формате ГГГГ-ММ-ДД:")


@router.message(ChallengeForm.start_date)
async def new_challenge_date(message: Message, state: FSMContext) -> None:
    try:
        start = date.fromisoformat((message.text or "").strip())
    except ValueError:
        await message.answer("Неверная дата. Пример: 2026-08-01")
        return
    await state.update_data(start_date=start.isoformat())
    await state.set_state(ChallengeForm.duration)
    await message.answer("Введите продолжительность в днях, например 21:")


@router.message(ChallengeForm.duration)
async def new_challenge_duration(message: Message, state: FSMContext) -> None:
    try:
        duration = int((message.text or "").strip())
        if not 1 <= duration <= 365:
            raise ValueError
    except ValueError:
        await message.answer("Введите целое число от 1 до 365.")
        return
    data = await state.get_data()
    start = date.fromisoformat(data["start_date"])
    end = start + timedelta(days=duration - 1)
    try:
        db.create_challenge(
            data["title"], start.isoformat(), end.isoformat(),
            RESULTS_CHAT_ID or message.chat.id,
        )
    except sqlite3.IntegrityError:
        await message.answer("Активный челлендж уже существует.")
        await state.clear()
        return
    await state.clear()
    await message.answer(
        f"✅ Создан «{escape(data['title'])}»\n"
        f"{start:%d.%m.%Y} — {end:%d.%m.%Y}",
        reply_markup=admin_keyboard(),
    )


@router.message(F.text == "👀 Кто не внёс")
async def missing_today(message: Message) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Нет активного челленджа.")
        return
    missing = db.get_missing_users(challenge.id, today_iso())
    text = "✅ Все внесли результат." if not missing else (
        "⏳ <b>Сегодня не внесли:</b>\n" +
        "\n".join(f"• {escape(row['full_name'])}" for row in missing)
    )
    await message.answer(text)


@router.message(F.text == "📥 Excel")
async def export_excel(message: Message) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Нет активного челленджа.")
        return
    content = create_excel(challenge)
    filename = f"challenge_{challenge.id}_{today_iso()}.xlsx"
    await message.answer_document(
        BufferedInputFile(content, filename=filename),
        caption="📊 Выгрузка результатов",
    )


@router.message(F.text == "🏆 Итоговые номинации")
async def nominations(message: Message) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Нет активного челленджа.")
        return
    await message.answer(build_nominations_text(challenge))


@router.message(F.text == "🏁 Завершить")
@router.message(Command("finish"))
async def finish(message: Message, bot: Bot) -> None:
    if not message.from_user or not is_admin(message.from_user.id):
        return
    challenge = db.get_active_challenge()
    if not challenge:
        await message.answer("Нет активного челленджа.")
        return
    final_text = build_final_text(challenge)
    chat_id = db.finish_challenge(challenge.id)
    await message.answer(final_text, reply_markup=main_keyboard(True))
    if chat_id and chat_id != message.chat.id:
        await bot.send_message(chat_id, final_text)


async def background_worker(bot: Bot) -> None:
    while True:
        try:
            now = now_local()
            challenge = db.get_active_challenge()
            if challenge:
                start = date.fromisoformat(challenge.start_date)
                end = date.fromisoformat(challenge.end_date)

                # Личные вечерние напоминания.
                reminder_key = f"reminder:{now.date().isoformat()}"
                if (
                    start <= now.date() <= end
                    and now.time() >= REMINDER_TIME
                    and not db.notification_was_sent(challenge.id, reminder_key)
                ):
                    for user in db.get_missing_users(
                        challenge.id, now.date().isoformat(), reminders_only=True
                    ):
                        try:
                            await bot.send_message(
                                user["telegram_id"],
                                "⏰ Вы ещё не внесли сегодняшний результат.",
                            )
                        except Exception:
                            logging.exception(
                                "Не удалось отправить напоминание пользователю %s",
                                user["telegram_id"],
                            )
                    db.mark_notification_sent(challenge.id, reminder_key)

                # Автопубликация рейтинга после 7-го, 14-го и 21-го дня.
                completed_days = (now.date() - start).days
                if completed_days in {7, 14, 21}:
                    week_number = completed_days // 7
                    weekly_key = f"weekly:{week_number}"
                    if not db.notification_was_sent(challenge.id, weekly_key):
                        reference = start + timedelta(days=completed_days - 1)
                        chat_id = challenge.results_chat_id
                        if chat_id:
                            await bot.send_message(
                                chat_id, build_weekly_text(challenge, reference)
                            )
                        db.mark_notification_sent(challenge.id, weekly_key)

                # Автоматическое завершение на следующий день.
                if now.date() > end:
                    final_key = "final"
                    if not db.notification_was_sent(challenge.id, final_key):
                        final_text = build_final_text(challenge)
                        chat_id = db.finish_challenge(challenge.id)
                        if chat_id:
                            await bot.send_message(chat_id, final_text)
                        db.mark_notification_sent(challenge.id, final_key)
        except Exception:
            logging.exception("Ошибка фонового процесса")
        await asyncio.sleep(60)


async def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    db.init()
    bot = Bot(
        token=BOT_TOKEN,
        default=DefaultBotProperties(parse_mode=ParseMode.HTML),
    )
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    await bot.set_my_commands(
        [
            BotCommand(command="start", description="Запустить бота"),
            BotCommand(command="rating", description="Общий рейтинг"),
            BotCommand(command="cancel", description="Отменить действие"),
            BotCommand(command="id", description="Показать Telegram ID"),
        ]
    )
    worker = asyncio.create_task(background_worker(bot))
    try:
        await dp.start_polling(bot)
    finally:
        worker.cancel()
        await bot.session.close()


if __name__ == "__main__":
    asyncio.run(main())
